"""
Supports parsing Dominion Excel (.xlsx) RCV result reports.
"""

import itertools
import logging

import openpyxl
from openpyxl.cell.cell import MergedCell

import rcvresults.parsers.common as common


_log = logging.getLogger(__name__)


def _iter_triples(iterator):
    # The "values" variable is the next triple to yield.
    # However, the first value yielded is a placeholder to start things
    # off and should be skipped over by the caller.
    values = ()
    while True:
        new_values = tuple(itertools.islice(iterator, 3))
        if not new_values:
            # Then yield the last triple of values, padding with Nones
            # as necessary to fill in missing values.
            missing = 3 - len(values)
            yield (*values, *(missing * [None]))
            break

        # Otherwise, "values" had three items, and we can iterate at least
        # one more time.
        yield values
        values = new_values


# TODO: move this to an utils.py module?
def iter_triples(iterator):
    """
    Yield the items in the given iterator as triples, padding the last item
    with None values as necessary.
    """
    iterator = _iter_triples(iterator)
    # Yield all items except the first placeholder value.
    yield from itertools.islice(iterator, 1, None)


def parse_sheet_1(wb, sheet_name):
    """
    Return a dict containing the contest name.
    """
    ws = wb[sheet_name]
    # The first cell of the first five rows has the following form:
    # 1: "Page: 1 / 2"
    # 2: None
    # 3: None
    # 4: "Final RCV Short Report\nCity and County of San Francisco\n"
    #    "November 8, 2022, Consolidated General Election"
    # 5: "PUBLIC DEFENDER"
    indexed_rows = enumerate(ws.iter_rows(values_only=True), start=1)
    for i, row in indexed_rows:
        value = row[0]
        if value is None:
            continue
        if 'Short Report' in value:
            break

    for i, row in indexed_rows:
        contest_name = row[0]
        break

    results = {
        'contest_name': contest_name,
    }
    return results


def iter_sheet2_rows(ws):
    """
    Yield the vote-total rows in "Sheet2" of the workbook.
    """
    indexed_rows = enumerate(ws.iter_rows(), start=1)
    # First, skip to the rows containing the candidates.
    for i, row in indexed_rows:
        cell = row[0]
        value = cell.value
        if value == 'Candidate':
            break

    i, row = next(indexed_rows)
    cell = row[0]
    value = cell.value
    assert value is None

    is_candidate = True
    for i, row in indexed_rows:
        cell = row[0]
        name = cell.value
        if name is None:
            break
        if name == 'Continuing Ballots Total':
            is_candidate = False

        yield (i, name, row, is_candidate)


def iter_sheet2_row(row):
    """
    Yield the values corresponding to the non-MergedCells in the given row.
    """
    for cell in row[1:]:
        if type(cell) is MergedCell:
            continue

        yield cell.value


def iter_row_rounds(row):
    """
    Yield the data in each round, one triple per round.
    """
    values = iter_sheet2_row(row)
    yield from iter_triples(values)


def parse_sheet2_row(row, is_candidate):
    rounds = []
    for data in iter_row_rounds(row):
        votes, percent, transfer = data
        round_data = {
            'votes': votes,
            'percent': percent,
            'transfer': transfer,
        }
        rounds.append(round_data)

    return rounds


def parse_sheet2(wb, sheet_name):
    ws = wb[sheet_name]

    candidates = []
    subtotals = []
    non_candidate_names = []
    rounds = {}
    for i, name, row, is_candidate in iter_sheet2_rows(ws):
        if is_candidate:
            candidates.append(name)
        else:
            name = common.get_subtotal_label(name)
            non_candidate_names.append(name)
        subtotals.append(name)
        row_rounds = parse_sheet2_row(row, is_candidate=is_candidate)
        rounds[name] = row_rounds

    # TODO: assert the contents of non_candidate_names?
    results = {
        'candidates': candidates,
        'non_candidate_names': non_candidate_names,
        'subtotals': subtotals,
        'rounds': rounds,
    }
    return results


def parse_excel_file(path):
    """
    Parse a workbook, and return a dict of results.
    """
    results = {}
    wb = openpyxl.load_workbook(filename=path)
    sheet_names = wb.sheetnames
    if sheet_names == ['Sheet1', 'Sheet2']:
        sheet_name_1, sheet_name_2 = sheet_names

    # In the 2019 election, the Excel files contained a single sheet that
    # was essentially the "Sheet1" above followed by "Sheet2" (so the same
    # formats of each, but combined into a single sheet).
    # These can be seen in the data-reports/2019-11-05 directory.
    elif sheet_names == ['RcvShortReport']:
        # Otherwise, use the same sheet for both.
        sheet_name_1, = sheet_names
        sheet_name_2 = sheet_name_1
    else:
        raise AssertionError(
            f'unexpected sheet names: {sheet_names}'
        )
    # Initialize the metadata dict with the contest name.
    metadata = parse_sheet_1(wb, sheet_name=sheet_name_1)
    results = parse_sheet2(wb, sheet_name=sheet_name_2)
    results['_metadata'] = metadata
    return results
